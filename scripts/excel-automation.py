#!/usr/bin/env python3
"""Parse compact SKU workbooks and fill Amazon category templates."""

from __future__ import annotations

import argparse
import base64
import json
import re
import sys
import unicodedata
from copy import copy
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qsl

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries


MAX_IMPORT_ROWS = 100
MAX_IMAGES = 10
URL_PATTERN = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)

SYSTEM_MANAGED_TEMPLATE_PREFIXES = (
    "contribution_sku",
    "product_type",
    "::record_action",
    "parentage_level[",
    "child_parent_sku_relationship[",
    "variation_theme",
    "item_name[",
    "item_name",
    "product_description[",
    "product_description",
    "bullet_point[",
    "generic_keyword[",
    "generic_keyword",
    "main_product_image_locator",
    "other_product_image_locator",
    "swatch_product_image_locator",
    "brand[",
    "brand_name",
    "manufacturer[",
    "manufacturer",
)


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


SKU_HEADERS = {
    "sku",
    "seller sku",
    "seller sku code",
    "ma sku",
    "sku code",
    "item sku",
}
MAIN_KEYWORD_HEADERS = {
    "main keyword",
    "primary keyword",
    "product name",
    "ten san pham",
    "ten san pham main keyword",
    "ten san pham primary keyword",
    "item name",
}
GENERIC_KEYWORD_HEADERS = {
    "generic keyword",
    "generic keywords",
    "generic search terms",
    "search terms",
    "genkey",
    "backend search terms",
}
IMAGE_HEADERS = {
    "link anh",
    "link anh trello",
    "anh trello",
    "trello",
    "trello link",
    "image",
    "images",
    "image link",
    "image links",
    "image url",
    "image urls",
    "main image url",
}


def header_kind(value: Any) -> str | None:
    text = normalize(value)
    if text in SKU_HEADERS:
        return "sku"
    if text in MAIN_KEYWORD_HEADERS:
        return "main_keyword"
    if text in GENERIC_KEYWORD_HEADERS:
        return "generic_keywords"
    if text in IMAGE_HEADERS or text.startswith("other image url"):
        return "image_urls"
    return None


def find_import_table(workbook: Any) -> tuple[Any, int, dict[str, list[int]]]:
    best: tuple[int, Any, int, dict[str, list[int]]] | None = None
    for worksheet in workbook.worksheets:
        for row_index in range(1, min(worksheet.max_row, 50) + 1):
            columns: dict[str, list[int]] = {}
            for column_index in range(1, min(worksheet.max_column, 400) + 1):
                kind = header_kind(worksheet.cell(row_index, column_index).value)
                if kind:
                    columns.setdefault(kind, []).append(column_index)
            required = sum(key in columns for key in ("sku", "main_keyword", "image_urls"))
            score = required * 10 + int("generic_keywords" in columns)
            if best is None or score > best[0]:
                best = (score, worksheet, row_index, columns)
            if required == 3 and "generic_keywords" in columns:
                return worksheet, row_index, columns
    if best and best[0] >= 30:
        return best[1], best[2], best[3]
    raise ValueError(
        "Không tìm thấy đủ cột SKU, Link ảnh (Trello) và Tên sản phẩm (Main Keyword)."
    )


def extract_urls(values: Iterable[tuple[Any, str | None]]) -> list[str]:
    urls: list[str] = []
    for value, hyperlink in values:
        candidates = []
        if hyperlink:
            candidates.append(hyperlink)
        candidates.extend(URL_PATTERN.findall(cell_text(value)))
        for candidate in candidates:
            url = candidate.rstrip(").,;]")
            if url not in urls:
                urls.append(url)
    return urls[:MAX_IMAGES]


def parse_workbook(source_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        source_path,
        read_only=False,
        data_only=True,
        keep_vba=source_path.suffix.lower() == ".xlsm",
    )
    worksheet, header_row, columns = find_import_table(workbook)
    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    seen_skus: set[str] = set()

    settings = cell_text(worksheet.cell(1, 1).value)
    if "dataRow=" in settings:
        attribute_row, _, _ = configured_template_rows(worksheet)
        template_columns = technical_columns(worksheet, attribute_row)
        parent_row, child_row = source_template_rows(worksheet, template_columns)
        data_start_row = min(parent_row, child_row)
    else:
        data_start_row = header_row + 1
    for row_index in range(data_start_row, worksheet.max_row + 1):
        sku = cell_text(worksheet.cell(row_index, columns["sku"][0]).value)
        main_keyword = cell_text(worksheet.cell(row_index, columns["main_keyword"][0]).value)
        generic_keywords = " | ".join(
            filter(
                None,
                (
                    cell_text(worksheet.cell(row_index, column).value)
                    for column in columns.get("generic_keywords", [])
                ),
            )
        )
        image_cells = []
        for column in columns["image_urls"]:
            cell = worksheet.cell(row_index, column)
            hyperlink = cell.hyperlink.target if cell.hyperlink else None
            image_cells.append((cell.value, hyperlink))
        image_urls = extract_urls(image_cells)

        if not any((sku, main_keyword, generic_keywords, image_urls)):
            continue
        missing = []
        if not sku:
            missing.append("SKU")
        if not main_keyword:
            missing.append("Tên sản phẩm (Main Keyword)")
        if not image_urls:
            missing.append("Link ảnh")
        if missing:
            raise ValueError(f"Dòng {row_index} thiếu: {', '.join(missing)}.")
        if sku.upper().endswith("_MAIN"):
            continue
        sku_key = sku.casefold()
        if sku_key in seen_skus:
            raise ValueError(f"SKU bị trùng ở dòng {row_index}: {sku}.")
        seen_skus.add(sku_key)
        if not generic_keywords:
            warnings.append(f"Dòng {row_index} ({sku}) chưa có Generic Keywords.")

        rows.append(
            {
                "source_row": row_index,
                "sku": sku,
                "main_keyword": main_keyword,
                "generic_keywords": generic_keywords,
                "image_urls": image_urls,
            }
        )
        if len(rows) >= MAX_IMPORT_ROWS:
            warnings.append(f"Chỉ đọc {MAX_IMPORT_ROWS} SKU đầu tiên.")
            break

    if not rows:
        raise ValueError("File Excel không có SKU hợp lệ.")
    return {
        "sheet_name": worksheet.title,
        "header_row": header_row,
        "rows": rows,
        "warnings": warnings,
    }


def create_sample(output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SKU Input"
    worksheet.append(
        [
            "SKU",
            "Link ảnh (Trello)",
            "Tên sản phẩm (Main Keyword)",
            "Generic Keywords",
        ]
    )
    sample_rows = [
        [
            "AOTT0731DL01C",
            "\n".join(
                [
                    "https://trello.com/1/cards/6a687525740c6bf7bc9fe6a5/attachments/6a74642c9a401d4b8bf125a6/download/AOTT0731DL01C_mk0.png",
                    "https://trello.com/1/cards/6a687525740c6bf7bc9fe6a5/attachments/6a746432c00e48ec9ec81805/download/AOTT0731DL01C_mk1.png",
                    "https://trello.com/1/cards/6a687525740c6bf7bc9fe6a5/attachments/6a746438c87599438c7d3d14/download/AOTT0731DL01C_mk2.png",
                ]
            ),
            "personalized golden retriever dog lover ornament",
            "golden retriever keepsake | dog owner christmas gift | personalized pet decor | holiday tree ornament",
        ],
        [
            "AOTT0731DL02C",
            "\n".join(
                [
                    "https://trello.com/1/cards/6a687525740c6bf7bc9fe6a5/attachments/6a74643d6f57586281295fc6/download/AOTT0731DL02C_mk0.png",
                    "https://trello.com/1/cards/6a687525740c6bf7bc9fe6a5/attachments/6a746443ff5d65eeccb059ac/download/AOTT0731DL02C_mk1.png",
                    "https://trello.com/1/cards/6a687525740c6bf7bc9fe6a5/attachments/6a7464498d3141c8a1de96b4/download/AOTT0731DL02C_mk2.png",
                ]
            ),
            "personalized labrador retriever dog lover ornament",
            "labrador retriever keepsake | dog owner christmas gift | personalized pet decor | holiday tree ornament",
        ],
        [
            "AOTT0731DL03C",
            "\n".join(
                [
                    "https://trello.com/1/cards/6a687525740c6bf7bc9fe6a5/attachments/6a74644f42f81be4c16a9069/download/AOTT0731DL03C_mk0.png",
                    "https://trello.com/1/cards/6a687525740c6bf7bc9fe6a5/attachments/6a7464545e3ec8a0e3fffb23/download/AOTT0731DL03C_mk1.png",
                ]
            ),
            "personalized french bulldog dog lover ornament",
            "french bulldog keepsake | dog owner christmas gift | personalized pet decor | holiday tree ornament",
        ],
    ]
    for sample_row in sample_rows:
        worksheet.append(sample_row)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:D{len(sample_rows) + 1}"
    for cell in worksheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 70
    worksheet.column_dimensions["C"].width = 42
    worksheet.column_dimensions["D"].width = 55
    for row_index in range(2, len(sample_rows) + 2):
        worksheet.row_dimensions[row_index].height = 54
        image_alignment = copy(worksheet.cell(row_index, 2).alignment)
        image_alignment.wrap_text = True
        image_alignment.vertical = "top"
        worksheet.cell(row_index, 2).alignment = image_alignment
    workbook.save(output_path)


def snapshot_row(worksheet: Any, row_index: int) -> list[dict[str, Any]]:
    snapshot = []
    for column_index in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row_index, column_index)
        snapshot.append(
            {
                "value": cell.value,
                "style": copy(cell._style),
                "hyperlink": copy(cell.hyperlink),
                "comment": copy(cell.comment),
            }
        )
    return snapshot


def restore_row(worksheet: Any, row_index: int, snapshot: list[dict[str, Any]]) -> None:
    for column_index, source in enumerate(snapshot, start=1):
        cell = worksheet.cell(row_index, column_index)
        cell.value = source["value"]
        cell._style = copy(source["style"])
        cell.hyperlink = copy(source["hyperlink"])
        cell.comment = copy(source["comment"])


def setting_row(worksheet: Any, name: str) -> int | None:
    settings = cell_text(worksheet.cell(1, 1).value)
    if settings.startswith("settings="):
        settings = settings[len("settings="):]
    match = re.search(rf"(?:^|&){name}=(\d+)(?:&|$)", settings)
    return int(match.group(1)) if match else None


def discover_attribute_row(worksheet: Any) -> int:
    best: tuple[int, int] | None = None
    markers = (
        "contribution_sku",
        "product_type",
        "item_name[",
        "product_description[",
        "generic_keyword[",
        "main_product_image_locator",
    )
    for row_index in range(1, min(worksheet.max_row, 100) + 1):
        names = [cell_text(worksheet.cell(row_index, column).value) for column in range(1, worksheet.max_column + 1)]
        score = sum(any(name.startswith(marker) for name in names) for marker in markers)
        if best is None or score > best[0]:
            best = (score, row_index)
    if not best or best[0] < 3:
        raise ValueError("Không tìm thấy dòng technical headers trong template Amazon.")
    return best[1]


def discover_label_row(worksheet: Any, attribute_row: int) -> int:
    configured = setting_row(worksheet, "labelRow")
    if configured is not None:
        return configured
    candidates = []
    for row_index in range(1, attribute_row):
        populated = sum(
            bool(cell_text(worksheet.cell(row_index, column).value))
            for column in range(1, worksheet.max_column + 1)
        )
        candidates.append((populated, row_index))
    return max(candidates)[1] if candidates else attribute_row


def configured_template_rows(worksheet: Any) -> tuple[int, int, int | None]:
    attribute_row = setting_row(worksheet, "attributeRow") or discover_attribute_row(worksheet)
    return attribute_row, discover_label_row(worksheet, attribute_row), setting_row(worksheet, "dataRow")


def amazon_template_settings(worksheet: Any) -> dict[str, str]:
    raw = cell_text(worksheet.cell(1, 1).value)
    if raw.startswith("settings="):
        raw = raw[len("settings="):]
    return {key: value for key, value in parse_qsl(raw, keep_blank_values=True)}


def decode_base64_text(value: str) -> str:
    if not value:
        return ""
    try:
        padded = value + "=" * (-len(value) % 4)
        return base64.urlsafe_b64decode(padded).decode("utf-8").strip()
    except (ValueError, UnicodeDecodeError):
        return ""


def format_product_type(value: str) -> str:
    words = re.sub(r"[_-]+", " ", cell_text(value)).strip().split()
    return " ".join(word.upper() if len(word) <= 3 else word.capitalize() for word in words)


def is_example_or_dummy_row(sku: str, title: str = "", brand: str = "") -> bool:
    s = sku.strip().casefold()
    t = title.strip().casefold()
    b = brand.strip().casefold()
    combined = f"{s} {t} {b}"

    # Exact or regex match for common dummy/sample SKUs (e.g. ABC123, ABC-123, XYZ123, SAMPLE_SKU)
    if re.search(r"^(abc|xyz|sample|example|test|demo|dummy|sku|placeholder|temp|none|n/a)[\d_-]*$", s):
        return True

    dummy_patterns = [
        "example",
        "sample",
        "demo",
        "test_sku",
        "test sku",
        "dummy",
        "abc123",
        "xyz123",
        "sku123",
        "[required]",
        "[optional]",
        "e.g.",
        "eg.",
    ]
    for pat in dummy_patterns:
        if pat in s:
            return True
    if s.startswith("ex_") or s.startswith("sample_") or s.startswith("test_"):
        return True
    if any(p in combined for p in ["example product", "example sku", "sample brand", "sample product", "example parent", "example child", "abc123", "dummy"]):
        return True
    return False


def scan_listing_template(template_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        template_path,
        read_only=False,
        data_only=False,
        keep_vba=template_path.suffix.lower() == ".xlsm",
    )
    if "Template" not in workbook.sheetnames:
        raise ValueError("Workbook không có sheet Template.")
    worksheet = workbook["Template"]
    attribute_row, label_row, data_row = configured_template_rows(worksheet)
    columns = technical_columns(worksheet, attribute_row)
    settings = amazon_template_settings(worksheet)

    contributor_id = settings.get("contributorId", "").strip()
    marketplace_id = settings.get("primaryMarketplaceId", "").strip()
    product_types: list[str] = []
    decoded_ptds = decode_base64_text(settings.get("ptds", ""))
    if decoded_ptds:
        try:
            parsed_ptds = json.loads(decoded_ptds)
            candidates = parsed_ptds if isinstance(parsed_ptds, list) else [parsed_ptds]
            product_types.extend(cell_text(value) for value in candidates if cell_text(value))
        except json.JSONDecodeError:
            product_types.append(decoded_ptds)

    decoded_browse = decode_base64_text(settings.get("browseClassifications", ""))
    if decoded_browse:
        try:
            for entry in json.loads(decoded_browse):
                product_type = cell_text(entry.get("productType")) if isinstance(entry, dict) else ""
                if product_type:
                    product_types.append(product_type)
        except (json.JSONDecodeError, TypeError):
            pass

    product_type_column = optional_column(columns, "product_type")
    if product_type_column is not None:
        for row_index in range(attribute_row + 1, min(worksheet.max_row, attribute_row + 50) + 1):
            product_type = cell_text(worksheet.cell(row_index, product_type_column).value)
            if product_type:
                product_types.append(product_type)

    product_type = next((value for value in product_types if value), "")
    if not product_type:
        raise ValueError("Không nhận diện được loại phôi từ metadata Amazon trong file.")
    if not contributor_id:
        raise ValueError("Không nhận diện được shop vì file không có contributorId của Seller Central.")

    store_name = (
        settings.get("storeName")
        or settings.get("merchantName")
        or settings.get("sellerName")
        or settings.get("accountName")
        or settings.get("brandName")
        or settings.get("brand")
        or settings.get("merchant_name")
        or settings.get("store_name")
        or ""
    ).strip()

    sku_col = None
    for name, idx in columns.items():
        if name.startswith("contribution_sku") or name == "item_sku":
            sku_col = idx
            break

    if not store_name:
        for prefix in ("brand[", "brand_name", "merchant_name", "store_name", "manufacturer["):
            for row_index in range(attribute_row + 1, min(worksheet.max_row, attribute_row + 30) + 1):
                sku_val = cell_text(worksheet.cell(row_index, sku_col).value) if sku_col else ""
                if is_example_or_dummy_row(sku_val):
                    continue
                val = first_value_with_prefix(worksheet, row_index, columns, prefix)
                if val and val.casefold() not in ("generic", "na", "n/a", "unknown") and not is_example_or_dummy_row(sku_val, brand=val):
                    store_name = val
                    break
            if store_name:
                break

    if not store_name:
        stem = template_path.stem.strip()
        if " - " in stem:
            candidate = stem.split(" - ")[0].strip()
            if candidate and not candidate.lower().startswith("amazon") and not candidate.lower().startswith("flat"):
                store_name = candidate
        elif "_" in stem:
            candidate = stem.split("_")[0].strip()
            if candidate and not candidate.lower().startswith("amazon") and not candidate.lower().startswith("flat"):
                store_name = candidate

    return {
        "sheet_name": worksheet.title,
        "attribute_row": attribute_row,
        "label_row": label_row,
        "data_row": data_row,
        "column_count": worksheet.max_column,
        "contributor_id": contributor_id,
        "shop_key": contributor_id.rsplit(".", 1)[-1],
        "marketplace_id": marketplace_id.rsplit(".", 1)[-1],
        "template_identifier": settings.get("templateIdentifier", "").strip(),
        "store_name": store_name,
        "product_type": product_type,
        "phoi_name": format_product_type(product_type),
    }


def technical_columns(worksheet: Any, attribute_row: int | None = None) -> dict[str, int]:
    if attribute_row is None:
        attribute_row = configured_template_rows(worksheet)[0]
    result: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        name = cell_text(worksheet.cell(attribute_row, column_index).value)
        if name:
            result[name] = column_index
    return result


def technical_header_signature(name: str) -> str:
    """Match the same Amazon field across nearby template revisions."""
    signature = cell_text(name).casefold()
    signature = re.sub(
        r"\[(marketplace_id|language_tag)=[^\]]+\]",
        lambda match: f"[{match.group(1)}=*]",
        signature,
    )
    return re.sub(r"\s+", "", signature)


def destination_column_for_source(
    source_name: str,
    destination_columns: dict[str, int],
    destination_signatures: dict[str, list[int]],
) -> int | None:
    if source_name in destination_columns:
        return destination_columns[source_name]
    candidates = destination_signatures.get(technical_header_signature(source_name), [])
    return candidates[0] if len(candidates) == 1 else None


def source_template_rows(worksheet: Any, columns: dict[str, int], data_row: int | None = None) -> tuple[int, int]:
    del data_row
    attribute_row = configured_template_rows(worksheet)[0]
    sku_column = find_column(columns, prefix="contribution_sku")
    parentage_column = optional_column(columns, "parentage_level[")
    relationship_column = optional_column(columns, "child_parent_sku_relationship[")

    title_col = None
    for name, idx in columns.items():
        if name.startswith("item_name[") or name == "item_name":
            title_col = idx
            break

    brand_col = None
    for name, idx in columns.items():
        if name.startswith("brand[") or name == "brand_name":
            brand_col = idx
            break

    parents: dict[str, list[int]] = {}
    children: list[tuple[int, str]] = []
    standalone_rows: list[tuple[int, bool]] = []

    for row_index in range(attribute_row + 1, worksheet.max_row + 1):
        sku = cell_text(worksheet.cell(row_index, sku_column).value)
        parentage = (
            cell_text(worksheet.cell(row_index, parentage_column).value).casefold()
            if parentage_column is not None
            else ""
        )
        if not sku:
            continue
        if parentage == "parent" or sku.upper().endswith("_MAIN"):
            parents.setdefault(sku.casefold(), []).append(row_index)
        else:
            parent_sku = (
                cell_text(worksheet.cell(row_index, relationship_column).value)
                if relationship_column is not None
                else ""
            )
            if parentage == "child" or parent_sku:
                children.append((row_index, parent_sku.casefold()))
            else:
                title = cell_text(worksheet.cell(row_index, title_col).value) if title_col else ""
                brand = cell_text(worksheet.cell(row_index, brand_col).value) if brand_col else ""
                standalone_rows.append(
                    (row_index, not is_example_or_dummy_row(sku, title, brand))
                )

    valid_pairs: list[tuple[int, int, bool]] = []
    for child_row, parent_sku in children:
        if parent_sku in parents:
            for parent_row in parents[parent_sku]:
                parent_sku_val = cell_text(worksheet.cell(parent_row, sku_column).value)
                child_sku_val = cell_text(worksheet.cell(child_row, sku_column).value)
                parent_title = cell_text(worksheet.cell(parent_row, title_col).value) if title_col else ""
                child_title = cell_text(worksheet.cell(child_row, title_col).value) if title_col else ""
                parent_brand = cell_text(worksheet.cell(parent_row, brand_col).value) if brand_col else ""

                is_p_dummy = is_example_or_dummy_row(parent_sku_val, parent_title, parent_brand)
                is_c_dummy = is_example_or_dummy_row(child_sku_val, child_title, parent_brand)
                is_real = not (is_p_dummy or is_c_dummy)

                valid_pairs.append((parent_row, child_row, is_real))

    # Priority 1: Real user product rows
    real_pairs = [p for p in valid_pairs if p[2]]
    if real_pairs:
        return real_pairs[0][0], real_pairs[0][1]

    real_standalone = [row for row, is_real in standalone_rows if is_real]
    if real_standalone:
        return real_standalone[0], real_standalone[0]

    # Fallback to example rows. This keeps Seller Central sample files usable.
    if valid_pairs:
        return valid_pairs[0][0], valid_pairs[0][1]
    if standalone_rows:
        return standalone_rows[0][0], standalone_rows[0][0]

    raise ValueError("Template cần có ít nhất một dòng mẫu listing lẻ.")


def values_with_prefix(
    worksheet: Any,
    row_index: int,
    columns: dict[str, int],
    prefix: str,
) -> list[str]:
    values = []
    for name, column_index in columns.items():
        if name.startswith(prefix):
            value = cell_text(worksheet.cell(row_index, column_index).value)
            if value and value not in values:
                values.append(value)
    return values


def first_value_with_prefix(
    worksheet: Any,
    row_index: int,
    columns: dict[str, int],
    prefix: str,
) -> str:
    values = values_with_prefix(worksheet, row_index, columns, prefix)
    return values[0] if values else ""


def dimension_text(worksheet: Any, row_index: int, columns: dict[str, int]) -> str:
    dimensions = []
    unit = ""
    for axis in ("width", "height", "depth"):
        value = ""
        for name, column_index in columns.items():
            if "item_depth_width_height" in name and f".{axis}.value" in name:
                value = cell_text(worksheet.cell(row_index, column_index).value)
            if "item_depth_width_height" in name and f".{axis}.unit" in name:
                unit = unit or cell_text(worksheet.cell(row_index, column_index).value)
        if value:
            dimensions.append(value)
    if not dimensions:
        return ""
    return f"{' x '.join(dimensions)} {unit}".strip()


def inspect_template(template_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        template_path,
        read_only=False,
        data_only=False,
        keep_vba=template_path.suffix.lower() == ".xlsm",
    )
    if "Template" not in workbook.sheetnames:
        raise ValueError("Workbook không có sheet Template.")
    worksheet = workbook["Template"]
    attribute_row, label_row, data_row = configured_template_rows(worksheet)
    columns = technical_columns(worksheet, attribute_row)
    required_prefixes = [
        "contribution_sku",
        "product_type",
        "item_name[",
        "product_description[",
        "generic_keyword[",
        "main_product_image_locator",
    ]
    for prefix in required_prefixes:
        find_column(columns, prefix=prefix)
    bullet_columns = []
    for occurrence in range(1, 6):
        marker = f"#{occurrence}."
        matching = [index for name, index in columns.items() if name.startswith("bullet_point[") and marker in name]
        if not matching:
            raise ValueError(f"Template thiếu Bullet Point #{occurrence}.")
        bullet_columns.append(get_column_letter(matching[0]))
    parent_row, child_row = source_template_rows(worksheet, columns, data_row)
    parent_sku = cell_text(worksheet.cell(parent_row, find_column(columns, prefix="contribution_sku")).value)
    child_sku = cell_text(worksheet.cell(child_row, find_column(columns, prefix="contribution_sku")).value)
    main_sku = (child_sku or parent_sku).replace("_MAIN", "").strip()
    color = first_value_with_prefix(worksheet, child_row, columns, "color[")
    if color.casefold() == child_sku.casefold():
        color = ""
    return {
        "main_sku": main_sku,
        "sheet_name": worksheet.title,
        "attribute_row": attribute_row,
        "label_row": label_row,
        "data_row": min(parent_row, child_row),
        "column_count": worksheet.max_column,
        "last_column": get_column_letter(worksheet.max_column),
        "source_parent_row": parent_row,
        "source_child_row": child_row,
        "listing_mode": "standalone" if parent_row == child_row else "legacy_parent_child",
        "product_type": first_value_with_prefix(worksheet, child_row, columns, "product_type"),
        "content_columns": {
            "sku": get_column_letter(find_column(columns, prefix="contribution_sku")),
            "title": get_column_letter(find_column(columns, prefix="item_name[")),
            "description": get_column_letter(find_column(columns, prefix="product_description[")),
            "bullet_points": bullet_columns,
            "generic_keywords": get_column_letter(find_column(columns, prefix="generic_keyword[")),
            "main_image": get_column_letter(find_column(columns, prefix="main_product_image_locator")),
        },
        "defaults": {
            "material": ", ".join(values_with_prefix(worksheet, child_row, columns, "material[")),
            "size_capacity": dimension_text(worksheet, child_row, columns),
            "color": color,
            "package_contents": ", ".join(values_with_prefix(worksheet, child_row, columns, "included_components[")),
            "features": values_with_prefix(worksheet, child_row, columns, "special_feature["),
            "country_of_origin": first_value_with_prefix(worksheet, child_row, columns, "country_of_origin["),
        },
    }


def extract_valid_options_for_column(
    workbook: Any,
    worksheet: Any,
    column_index: int,
    column_name: str,
    product_type: str = "",
) -> list[str]:
    def values_from_range(sheet: Any, coordinate: str) -> list[str]:
        found: list[str] = []
        cells = sheet[coordinate.replace("$", "").strip()]
        if not isinstance(cells, tuple):
            cells = ((cells,),)
        elif cells and not isinstance(cells[0], tuple):
            cells = (cells,)
        for row in cells:
            for cell in row:
                value = cell_text(cell.value).strip()
                if value and value not in found and not value.startswith("="):
                    found.append(value)
        return found

    def values_from_defined_name(name: str) -> list[str]:
        if not name or not hasattr(workbook, "defined_names") or name not in workbook.defined_names:
            return []
        try:
            definition = workbook.defined_names[name]
            found: list[str] = []
            for title, coordinate in definition.destinations:
                if title in workbook.sheetnames:
                    for value in values_from_range(workbook[title], coordinate):
                        if value not in found:
                            found.append(value)
            return found
        except Exception:
            return []

    def formula_options(formula: str) -> list[str]:
        formula = formula.strip().lstrip("=")
        if formula.startswith('"') and formula.endswith('"'):
            return [value.strip() for value in formula[1:-1].split(",") if value.strip()]
        direct = values_from_defined_name(formula)
        if direct:
            return direct
        if "!" in formula and not formula.upper().startswith(("IF(", "INDIRECT(")):
            try:
                sheet_ref, coordinate = formula.split("!", 1)
                sheet_name = sheet_ref.strip("'\"").strip()
                if sheet_name in workbook.sheetnames:
                    return values_from_range(workbook[sheet_name], coordinate)
            except Exception:
                pass

        # Amazon builds validation names as PRODUCT_TYPE + technical header
        # with brackets, # and = removed. Resolve it without evaluating Excel.
        if product_type:
            product_key = re.sub(r"\s+", "", product_type.replace("-", "_"))
            if product_key[:1].isdigit():
                product_key = f"_{product_key}"
            suffix = re.sub(r"[\[\]#=]", "", column_name).replace("-", "_").replace(" ", "")
            computed = values_from_defined_name(f"{product_key}{suffix}")
            if computed:
                return computed
        return []

    options: list[str] = []
    try:
        if hasattr(worksheet, "data_validations") and worksheet.data_validations:
            for dv in worksheet.data_validations.dataValidation:
                if dv.type != "list":
                    continue
                in_range = False
                for r in dv.ranges:
                    if r.min_col <= column_index <= r.max_col:
                        in_range = True
                        break
                if in_range and dv.formula1:
                    options = formula_options(str(dv.formula1))
                    if options:
                        return options[:1000]
    except Exception:
        pass

    for sheet_candidate in ["Valid Values", "ValidValues", "Data Definitions"]:
        if sheet_candidate in workbook.sheetnames:
            try:
                vv_sheet = workbook[sheet_candidate]
                target_col = None
                for row_idx in range(1, min(6, vv_sheet.max_row + 1)):
                    for col_idx in range(1, vv_sheet.max_column + 1):
                        header_val = cell_text(vv_sheet.cell(row_idx, col_idx).value).casefold()
                        if (
                            "merchant_shipping_group_name" in header_val
                            or ("shipping" in header_val and ("template" in header_val or "group" in header_val or "merchant" in header_val))
                            or column_name.casefold() in header_val
                        ):
                            target_col = col_idx
                            break
                    if target_col:
                        break

                if target_col:
                    for r in range(2, vv_sheet.max_row + 1):
                        val = cell_text(vv_sheet.cell(r, target_col).value).strip()
                        if val and val not in options and not val.startswith("=") and not val.casefold().startswith("example"):
                            options.append(val)
                    if options:
                        return options
            except Exception:
                pass

    return options


def is_system_managed_template_column(column_name: str) -> bool:
    return column_name.startswith(SYSTEM_MANAGED_TEMPLATE_PREFIXES)


def blank_template_setup_field(column_name: str, label: str) -> tuple[str, str] | None:
    """Return the one-time setup field for a truly blank Amazon template."""
    norm_col = column_name.lower().replace("-", "_")
    norm_label = normalize(label)

    # Additional occurrences remain available when Excel actually renders
    # them red, but the one-time blank setup asks for only the primary value.
    if re.search(r'#(?:[2-9]|\d{2,})\b', norm_col):
        return None

    if (
        any(norm_col.startswith(prefix) for prefix in (
            "material[",
            "material_type[",
            "outer_material_type[",
            "material_composition[",
        ))
        or norm_col in {"material", "material_type", "outer_material_type"}
        or norm_label in {"material", "material type", "chat lieu", "vat lieu"}
    ):
        return "material", "Material"

    if (
        norm_col.startswith(("fabric_type[", "fabric_type_"))
        or norm_col == "fabric_type"
        or "fabric type" in norm_label
        or "loai vai" in norm_label
    ):
        return "fabric_type", "Fabric Type"

    if (
        norm_col.startswith(("color[", "color_name[", "colour[", "colour_name["))
        or norm_col in {"color", "color_name", "colour", "colour_name"}
        or norm_label in {"color", "color name", "colour", "colour name", "mau sac", "mau"}
    ):
        return "color", "Color"

    if (
        norm_col.startswith(("item_shape[", "shape["))
        or norm_col in {"item_shape", "shape"}
        or norm_label in {"item shape", "shape", "hinh dang"}
    ):
        return "item_shape", "Item Shape"

    package_dimensions = (
        (
            "package_length",
            "package_dimensions[#1.length",
            "package_depth_width_height[#1.length.value]",
            "item_package_length",
        ),
        (
            "package_width",
            "package_dimensions[#1.width",
            "package_depth_width_height[#1.width.value]",
            "item_package_width",
        ),
        (
            "package_height",
            "package_dimensions[#1.height",
            "package_depth_width_height[#1.height.value]",
            "item_package_height",
        ),
    )
    for key, patterns, display_label in (
        ("package_length", package_dimensions[0], "Package Length (inches)"),
        ("package_width", package_dimensions[1], "Package Width (inches)"),
        ("package_height", package_dimensions[2], "Package Height (inches)"),
    ):
        label_phrase = key.replace("_", " ")
        if any(pattern in norm_col for pattern in patterns) or label_phrase in norm_label:
            if (
                not norm_col.endswith(".unit")
                and "unit_of_measure" not in norm_col
                and "unit" not in norm_label
            ):
                return key, display_label

    return None


def required_template_fields(template_path: Path, payload_path: Path | None = None) -> dict[str, Any]:
    keep_vba = template_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(template_path, read_only=False, data_only=False, keep_vba=keep_vba)
    if "Template" not in workbook.sheetnames:
        raise ValueError("Workbook không có sheet Template.")
    worksheet = workbook["Template"]
    attribute_row, label_row, data_row = configured_template_rows(worksheet)
    columns = technical_columns(worksheet, attribute_row)
    scan = scan_listing_template(template_path)
    target_row = data_row or attribute_row + 1

    payload_data: dict[str, Any] = {}
    field_values: dict[str, str] = {}
    if payload_path is not None and payload_path.exists():
        try:
            parsed_payload = json.loads(payload_path.read_text(encoding="utf-8"))
            if isinstance(parsed_payload, dict):
                payload_data = parsed_payload
                field_values = parsed_payload.get("field_values", {})
        except Exception:
            pass
    include_setup_fields = payload_data.get("include_setup_fields") is True
    show_existing_setup_fields = payload_data.get("show_existing_setup_fields") is True

    if field_values:
        for column_name, val in field_values.items():
            if column_name in columns and val is not None and str(val).strip():
                worksheet.cell(target_row, columns[column_name]).value = str(val).strip()

    def color_is_red(color: Any) -> bool:
        if color is None or getattr(color, "type", None) != "rgb":
            return False
        rgb = cell_text(getattr(color, "rgb", "")).upper()[-6:]
        if len(rgb) != 6:
            return False
        try:
            red, green, blue = int(rgb[:2], 16), int(rgb[2:4], 16), int(rgb[4:], 16)
            return red >= 180 and green <= 100 and blue <= 100
        except ValueError:
            return False

    def color_is_white(color: Any) -> bool:
        if color is None or getattr(color, "type", None) != "rgb":
            return False
        return cell_text(getattr(color, "rgb", "")).upper()[-6:] == "FFFFFF"

    def border_has_red(border: Any) -> bool:
        return bool(border) and any(
            color_is_red(getattr(getattr(border, side_name, None), "color", None))
            for side_name in ("left", "right", "top", "bottom")
        )

    def fill_is_white(fill: Any, allow_unset: bool = False) -> bool:
        if not fill or getattr(fill, "fill_type", None) is None:
            return allow_unset
        return color_is_white(getattr(fill, "fgColor", None))

    def border_has_visible_style(border: Any) -> bool:
        return bool(border) and any(
            getattr(getattr(border, side_name, None), "style", None)
            or getattr(getattr(border, side_name, None), "color", None)
            for side_name in ("left", "right", "top", "bottom")
        )

    def fill_has_visible_style(fill: Any) -> bool:
        return bool(fill) and getattr(fill, "fill_type", None) is not None

    def rule_style(rule: Any) -> tuple[Any, Any]:
        differential_style = getattr(rule, "dxf", None)
        if differential_style is not None:
            return (
                getattr(differential_style, "border", None),
                getattr(differential_style, "fill", None),
            )
        return getattr(rule, "border", None), getattr(rule, "fill", None)

    product_type_column = optional_column(columns, "product_type")
    effective_product_type = (
        cell_text(worksheet.cell(target_row, product_type_column).value)
        if product_type_column is not None
        else ""
    ) or cell_text(scan.get("product_type"))

    token_pattern = re.compile(
        r'''\s*(?:
            (?P<string>"(?:[^"]|"")*")
          | (?P<reference>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ ]*)!\$?[A-Z]{1,3}\$?\d+)
          | (?P<cell>\$?[A-Z]{1,3}\$?\d+)
          | (?P<number>\d+(?:\.\d+)?)
          | (?P<operator><>|>=|<=|=|>|<)
          | (?P<lparen>\()
          | (?P<rparen>\))
          | (?P<comma>,)
          | (?P<identifier>[A-Za-z_][A-Za-z0-9_.]*)
        )''',
        re.VERBOSE,
    )
    name_cache: dict[str, Any] = {}

    def named_range_values(name: str) -> list[str]:
        definition = workbook.defined_names.get(name)
        if definition is None:
            return []
        values: list[str] = []
        try:
            for sheet_name, coordinate in definition.destinations:
                if sheet_name not in workbook.sheetnames:
                    continue
                cells = workbook[sheet_name][coordinate.replace("$", "")]
                if not isinstance(cells, tuple):
                    cells = ((cells,),)
                elif cells and not isinstance(cells[0], tuple):
                    cells = (cells,)
                for row in cells:
                    for cell in row:
                        value = cell_text(cell.value)
                        if value:
                            values.append(value)
        except Exception:
            return []
        return values

    def excel_truthy(value: Any) -> bool:
        if isinstance(value, str):
            return bool(value)
        return bool(value)

    def excel_equal(left: Any, right: Any) -> bool:
        if isinstance(left, str) or isinstance(right, str):
            return cell_text(left).casefold() == cell_text(right).casefold()
        return left == right

    def resolve_cell_reference(reference: str, named_formula: bool) -> Any:
        match = re.search(r"!\$?([A-Z]{1,3})(\$?)(\d+)$", reference)
        if match:
            column_name, absolute_row, row_text = match.groups()
            row_index = int(row_text) if absolute_row else target_row
            sheet_name = reference.split("!", 1)[0].strip("'")
            target_sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else worksheet
        else:
            direct = re.fullmatch(r"\$?([A-Z]{1,3})(\$?)(\d+)", reference)
            if direct is None:
                raise ValueError(f"Tham chiếu Excel không hỗ trợ: {reference}")
            column_name, _, row_text = direct.groups()
            row_index = int(row_text)
            target_sheet = worksheet
        column_index = column_index_from_string(column_name)
        if target_sheet == worksheet and row_index == target_row and column_index == product_type_column:
            return effective_product_type
        return target_sheet.cell(row_index, column_index).value or ""

    class FormulaParser:
        def __init__(self, formula: str, named_formula: bool = False):
            self.formula = formula.strip().lstrip("=")
            self.named_formula = named_formula
            self.tokens: list[tuple[str, str]] = []
            self.position = 0
            cursor = 0
            for match in token_pattern.finditer(self.formula):
                if self.formula[cursor:match.start()].strip():
                    raise ValueError(f"Công thức Excel không hỗ trợ: {self.formula}")
                kind = next(key for key, value in match.groupdict().items() if value is not None)
                self.tokens.append((kind, match.group(kind)))
                cursor = match.end()
            if self.formula[cursor:].strip():
                raise ValueError(f"Công thức Excel không hỗ trợ: {self.formula}")

        def peek(self, kind: str | None = None) -> tuple[str, str] | None:
            if self.position >= len(self.tokens):
                return None
            token = self.tokens[self.position]
            return token if kind is None or token[0] == kind else None

        def take(self, kind: str | None = None) -> tuple[str, str]:
            token = self.peek(kind)
            if token is None:
                raise ValueError(f"Công thức Excel không hợp lệ: {self.formula}")
            self.position += 1
            return token

        def parse(self) -> Any:
            value = self.expression()
            if self.peek() is not None:
                raise ValueError(f"Công thức Excel còn dữ liệu chưa xử lý: {self.formula}")
            return value

        def expression(self) -> Any:
            left = self.primary()
            operator = self.peek("operator")
            if operator is None:
                return left
            self.take("operator")
            right = self.primary()
            op = operator[1]
            if op == "=":
                return excel_equal(left, right)
            if op == "<>":
                return not excel_equal(left, right)
            try:
                left_number, right_number = float(left), float(right)
            except (TypeError, ValueError):
                left_number, right_number = cell_text(left).casefold(), cell_text(right).casefold()
            return {
                ">": left_number > right_number,
                "<": left_number < right_number,
                ">=": left_number >= right_number,
                "<=": left_number <= right_number,
            }[op]

        def primary(self) -> Any:
            token = self.take()
            kind, value = token
            if kind == "string":
                return value[1:-1].replace('""', '"')
            if kind == "number":
                return float(value) if "." in value else int(value)
            if kind in {"reference", "cell"}:
                return resolve_cell_reference(value, self.named_formula)
            if kind == "lparen":
                result = self.expression()
                self.take("rparen")
                return result
            if kind != "identifier":
                raise ValueError(f"Token Excel không hỗ trợ: {value}")
            if self.peek("lparen") is not None:
                self.take("lparen")
                arguments: list[Any] = []
                if self.peek("rparen") is None:
                    while True:
                        arguments.append(self.expression())
                        if self.peek("comma") is None:
                            break
                        self.take("comma")
                self.take("rparen")
                function = value.upper()
                if function == "AND":
                    return all(excel_truthy(argument) for argument in arguments)
                if function == "OR":
                    return any(excel_truthy(argument) for argument in arguments)
                if function == "NOT":
                    return not excel_truthy(arguments[0])
                if function == "LEN":
                    return len(cell_text(arguments[0]))
                if function == "IF":
                    return arguments[1] if excel_truthy(arguments[0]) else arguments[2]
                if function == "COUNTIF":
                    values = arguments[0] if isinstance(arguments[0], list) else []
                    return sum(excel_equal(item, arguments[1]) for item in values)
                raise ValueError(f"Hàm Excel không hỗ trợ: {function}")
            upper_value = value.upper()
            if upper_value == "TRUE":
                return True
            if upper_value == "FALSE":
                return False
            cached = name_cache.get(value.casefold(), None)
            if cached is not None:
                return cached
            definition = workbook.defined_names.get(value)
            if definition is None:
                raise ValueError(f"Không tìm thấy Excel defined name: {value}")
            definition_text = cell_text(getattr(definition, "attr_text", ""))
            if definition_text.startswith("'") and "!" in definition_text and not any(
                marker in definition_text for marker in ("=", "<>", "(")
            ):
                result: Any = named_range_values(value)
            else:
                result = FormulaParser(definition_text, named_formula=True).parse()
            name_cache[value.casefold()] = result
            return result

    def evaluate_rule(rule: Any) -> bool | None:
        formula = " ".join(rule.formula or []).strip()
        try:
            return excel_truthy(FormulaParser(formula).parse())
        except Exception:
            # openpyxl does not calculate Conditional Formatting. If Amazon
            # introduces a formula we cannot evaluate, do not invent a red
            # field that Excel may actually render as gray/non-applicable.
            return None

    product_type_column = optional_column(columns, "product_type")
    effective_product_type = (
        cell_text(worksheet.cell(target_row, product_type_column).value)
        if product_type_column is not None
        else ""
    ) or cell_text(scan.get("product_type"))

    brand_name = cell_text(payload_data.get("brand_name")) or "Generic"

    if product_type_column is not None and effective_product_type:
        worksheet.cell(target_row, product_type_column).value = effective_product_type
    set_optional_value(worksheet, target_row, columns, brand_name, prefix="brand[")
    set_optional_value(worksheet, target_row, columns, brand_name, prefix="manufacturer[")
    set_optional_value(worksheet, target_row, columns, "Sample Title", prefix="item_name")
    set_optional_value(worksheet, target_row, columns, "SAMPLE_SKU", prefix="contribution_sku")
    set_optional_value(worksheet, target_row, columns, "Sample Description", prefix="product_description")
    for occurrence in range(1, 6):
        try:
            set_repeated_value(
                worksheet,
                target_row,
                columns,
                "bullet_point[",
                occurrence,
                f"Sample Bullet {occurrence}",
            )
        except Exception:
            pass
    set_optional_value(worksheet, target_row, columns, "sample keywords", prefix="generic_keyword")

    record_action_column = optional_column(columns, "::record_action")
    if record_action_column is not None and not cell_text(worksheet.cell(target_row, record_action_column).value):
        worksheet.cell(target_row, record_action_column).value = "Create or Replace"

    field_values = payload_data.get("field_values", {})
    if isinstance(field_values, dict):
        for column_name, val in field_values.items():
            if column_name in columns and val is not None and str(val).strip():
                worksheet.cell(target_row, columns[column_name]).value = str(val).strip()

    # Calculate the effective style for every empty cell. Conditional
    # Formatting has priority over the stored/base style, so a base red border
    # must not leak through when Excel currently renders a gray rule.
    active_red_columns: set[int] = set()
    conditional_rules_by_column: dict[int, list[Any]] = {}
    if hasattr(worksheet, "conditional_formatting"):
        for conditional_range in worksheet.conditional_formatting:
            try:
                for coordinate in str(conditional_range.sqref).split():
                    min_col, min_row, max_col, max_row = range_boundaries(coordinate)
                    if min_row <= target_row <= max_row:
                        for column_index in range(min_col, max_col + 1):
                            conditional_rules_by_column.setdefault(column_index, []).extend(
                                worksheet.conditional_formatting[conditional_range]
                            )
            except Exception:
                continue

    for column_index in range(1, worksheet.max_column + 1):
        target_cell = worksheet.cell(target_row, column_index)
        if cell_text(target_cell.value):
            continue

        effective_border = getattr(target_cell, "border", None)
        effective_fill = getattr(target_cell, "fill", None)
        border_resolved = False
        fill_resolved = False
        style_is_unknown = False

        for rule in sorted(
            conditional_rules_by_column.get(column_index, []),
            key=lambda candidate: candidate.priority,
        ):
            applies = evaluate_rule(rule)
            if applies is None:
                style_is_unknown = True
                break
            if not applies:
                continue

            rule_border, rule_fill = rule_style(rule)
            if not border_resolved and border_has_visible_style(rule_border):
                effective_border = rule_border
                border_resolved = True
            if not fill_resolved and fill_has_visible_style(rule_fill):
                effective_fill = rule_fill
                fill_resolved = True
            if rule.stopIfTrue:
                break

        if style_is_unknown:
            continue
        if (
            border_has_red(effective_border)
            and fill_is_white(effective_fill, allow_unset=True)
        ):
            active_red_columns.add(column_index)

    fields: list[dict[str, Any]] = []
    managed_count = 0
    added_blank_setup_fields: set[str] = set()
    for column_name, column_index in sorted(columns.items(), key=lambda entry: entry[1]):
        if is_system_managed_template_column(column_name):
            managed_count += 1
            continue

        raw_label = cell_text(worksheet.cell(label_row, column_index).value) or column_name
        setup_field = blank_template_setup_field(column_name, raw_label) if include_setup_fields else None
        is_blank_setup_field = bool(
            setup_field and setup_field[0] not in added_blank_setup_fields
        )

        existing_value = cell_text(worksheet.cell(target_row, column_index).value)
        if existing_value and not (is_blank_setup_field and show_existing_setup_fields):
            continue

        if column_index not in active_red_columns and not is_blank_setup_field:
            continue

        options = extract_valid_options_for_column(
            workbook,
            worksheet,
            column_index,
            column_name,
            scan["product_type"],
        )

        fields.append({
            "id": column_name,
            "technical_name": column_name,
            "label": setup_field[1] if is_blank_setup_field else raw_label,
            "column": get_column_letter(column_index),
            "input_type": "select" if options else "text",
            "options": options,
            "default_value": existing_value if is_blank_setup_field else "",
        })
        if is_blank_setup_field and setup_field:
            added_blank_setup_fields.add(setup_field[0])

    return {
        **scan,
        "required_fields": fields,
        "managed_fields_count": managed_count,
    }


def prepare_standalone_template(
    template_path: Path,
    payload_path: Path,
    output_path: Path,
) -> dict[str, Any]:
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    field_values = payload.get("field_values", {})
    if not isinstance(field_values, dict):
        raise ValueError("Dữ liệu các trường template không hợp lệ.")
    brand_name = cell_text(payload.get("brand_name")) or "Generic"

    keep_vba = template_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(template_path, read_only=False, data_only=False, keep_vba=keep_vba)
    worksheet = workbook["Template"]
    attribute_row, _, data_row = configured_template_rows(worksheet)
    columns = technical_columns(worksheet, attribute_row)
    target_row = data_row or attribute_row + 1
    standalone_snapshot = snapshot_row(worksheet, target_row)
    scan = scan_listing_template(template_path)

    # 1. Normalize supplied field values
    normalized_values: dict[str, str] = {}
    for col_id, val in field_values.items():
        val_str = cell_text(val)
        if not val_str:
            continue
        col_idx = columns.get(col_id)
        if col_idx is not None:
            options = extract_valid_options_for_column(workbook, worksheet, col_idx, col_id, scan["product_type"])
            if options:
                matched = next((opt for opt in options if opt.casefold() == val_str.casefold()), None)
                if matched is not None:
                    val_str = matched
        normalized_values[col_id] = val_str

    # 2. Re-scan template with current field_values to check for any remaining red cells
    check_scan = required_template_fields(template_path, payload_path)
    remaining_missing = check_scan["required_fields"]

    if remaining_missing:
        missing_labels = [f["label"] for f in remaining_missing]
        missing_str = ", ".join(f'"{name}"' for name in missing_labels[:5])
        more = f" và {len(missing_labels) - 5} trường khác" if len(missing_labels) > 5 else ""
        raise ValueError(f'Còn ô đỏ cần điền nốt trước khi lưu: {missing_str}{more}.')

    keep_vba = template_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(template_path, read_only=False, data_only=False, keep_vba=keep_vba)
    worksheet = workbook["Template"]
    attribute_row, _, data_row = configured_template_rows(worksheet)
    columns = technical_columns(worksheet, attribute_row)
    target_row = data_row or attribute_row + 1
    standalone_snapshot = snapshot_row(worksheet, target_row)

    record_action_column = optional_column(columns, "::record_action")
    record_action = ""
    if record_action_column is not None:
        record_options = extract_valid_options_for_column(
            workbook,
            worksheet,
            record_action_column,
            "::record_action",
            scan["product_type"],
        )
        record_action = next(
            (value for value in record_options if "default" in value.casefold() or "create or replace" in value.casefold()),
            record_options[0] if record_options else "",
        )

    for row_index in range(attribute_row + 1, worksheet.max_row + 1):
        for column_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row_index, column_index)
            if not isinstance(cell, MergedCell):
                cell.value = None
                cell.hyperlink = None
                cell.comment = None

    # Preserve defaults already configured on the one standalone source row.
    # Only missing red cells are collected by the upload UI, so deleting this
    # row would turn valid non-red defaults into new red cells after saving.
    restore_row(worksheet, target_row, standalone_snapshot)

    set_value(worksheet, target_row, columns, "STANDALONE_TEMPLATE", prefix="contribution_sku")
    set_value(worksheet, target_row, columns, scan["product_type"], prefix="product_type")
    if record_action_column is not None and record_action:
        worksheet.cell(target_row, record_action_column).value = record_action
    set_optional_value(worksheet, target_row, columns, "", prefix="parentage_level[")
    set_optional_value(worksheet, target_row, columns, "", prefix="child_parent_sku_relationship[")
    set_optional_value(worksheet, target_row, columns, "", prefix="variation_theme")
    set_value(worksheet, target_row, columns, "Standalone template title", prefix="item_name")
    set_value(worksheet, target_row, columns, "Standalone template description", prefix="product_description")
    for occurrence in range(1, 6):
        set_repeated_value(
            worksheet,
            target_row,
            columns,
            "bullet_point[",
            occurrence,
            f"Standalone template bullet {occurrence}",
        )
    set_value(worksheet, target_row, columns, "standalone template keywords", prefix="generic_keyword")
    set_optional_value(worksheet, target_row, columns, brand_name, prefix="brand[")
    set_optional_value(worksheet, target_row, columns, brand_name, prefix="manufacturer[")

    for column_name, value in normalized_values.items():
        column_index = columns.get(column_name)
        if column_index is not None:
            worksheet.cell(target_row, column_index).value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    # STRICT POST-SAVE VERIFICATION: Scan the saved workbook to verify 0 red cells remain
    post_scan = required_template_fields(output_path)
    if post_scan.get("required_fields"):
        missing_labels = [f["label"] for f in post_scan["required_fields"]]
        missing_str = ", ".join(f'"{name}"' for name in missing_labels[:5])
        more = f" và {len(missing_labels) - 5} trường khác" if len(missing_labels) > 5 else ""
        error_json = json.dumps(post_scan["required_fields"])
        raise ValueError(f'Còn ô đỏ chưa điền trong file đã lưu: {missing_str}{more}. Vui lòng bổ sung trước khi lưu.__REQUIRED_FIELDS_JSON__{error_json}')

    inspection = inspect_template(output_path)
    inspection["is_blank"] = False
    inspection["is_ready"] = True
    inspection["listing_mode"] = "standalone"
    return inspection


def validate_shipping_for_row(workbook: Any, worksheet: Any, row_index: int, columns: dict[str, int]) -> None:
    shipping_col = optional_column(columns, "merchant_shipping_group_name")
    if shipping_col is None:
        for name, col_idx in columns.items():
            if "shipping" in name.lower() and ("merchant" in name.lower() or "group" in name.lower() or "template" in name.lower()):
                shipping_col = col_idx
                break
    if shipping_col is not None:
        current_val = cell_text(worksheet.cell(row_index, shipping_col).value).strip()
        options = extract_valid_options_for_column(workbook, worksheet, shipping_col, "merchant_shipping_group_name")
        if options:
            matched = next((opt for opt in options if opt.casefold() == current_val.casefold()), None)
            if matched:
                worksheet.cell(row_index, shipping_col).value = matched
            else:
                worksheet.cell(row_index, shipping_col).value = options[0]


def map_template_to_blank(
    source_path: Path,
    destination_path: Path,
    output_path: Path,
    target_brand: str = "",
) -> dict[str, Any]:
    """Copy row values by Amazon technical header while retaining destination workbook metadata."""
    source_workbook = load_workbook(
        source_path,
        read_only=False,
        data_only=False,
        keep_vba=source_path.suffix.lower() == ".xlsm",
    )
    destination_workbook = load_workbook(
        destination_path,
        read_only=False,
        data_only=False,
        keep_vba=destination_path.suffix.lower() == ".xlsm",
    )
    if "Template" not in source_workbook.sheetnames:
        raise ValueError("Template nguồn không có sheet Template.")
    if "Template" not in destination_workbook.sheetnames:
        raise ValueError("Blank template của shop đích không có sheet Template.")

    source_sheet = source_workbook["Template"]
    destination_sheet = destination_workbook["Template"]
    source_attribute_row, _, source_data_row = configured_template_rows(source_sheet)
    destination_attribute_row, _, destination_data_row = configured_template_rows(destination_sheet)
    source_columns = technical_columns(source_sheet, source_attribute_row)
    destination_columns = technical_columns(destination_sheet, destination_attribute_row)
    source_parent_row, source_child_row = source_template_rows(
        source_sheet,
        source_columns,
        source_data_row,
    )
    source_rows = [source_parent_row]
    if source_child_row != source_parent_row:
        source_rows.append(source_child_row)

    destination_signatures: dict[str, list[int]] = {}
    for name, column_index in destination_columns.items():
        destination_signatures.setdefault(technical_header_signature(name), []).append(column_index)

    target_start_row = destination_data_row or destination_attribute_row + 1
    if target_start_row <= destination_attribute_row:
        raise ValueError("Dòng bắt đầu dữ liệu của blank template shop đích không hợp lệ.")

    mapped_headers: set[str] = set()
    unmapped_headers: set[str] = set()
    source_value_count = 0
    mapped_value_count = 0
    for row_offset, source_row_index in enumerate(source_rows):
        destination_row_index = target_start_row + row_offset
        for source_name, source_column_index in source_columns.items():
            source_cell = source_sheet.cell(source_row_index, source_column_index)
            if source_cell.value is None and source_cell.hyperlink is None:
                continue
            if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
                # Formulas can contain workbook links or references. Mapping copies plain content only.
                continue
            source_value_count += 1
            destination_column_index = destination_column_for_source(
                source_name,
                destination_columns,
                destination_signatures,
            )
            if destination_column_index is None:
                unmapped_headers.add(source_name)
                continue

            val_to_set = source_cell.value
            if "shipping" in source_name.lower():
                options = extract_valid_options_for_column(
                    destination_workbook,
                    destination_sheet,
                    destination_column_index,
                    source_name,
                )
                if options:
                    matched = next(
                        (opt for opt in options if opt.casefold() == str(val_to_set).strip().casefold()),
                        None,
                    )
                    if matched:
                        val_to_set = matched
                    else:
                        val_to_set = options[0]

            destination_cell = destination_sheet.cell(destination_row_index, destination_column_index)
            destination_cell.value = val_to_set
            destination_cell.hyperlink = source_cell.hyperlink.target if source_cell.hyperlink else None
            mapped_headers.add(source_name)
            mapped_value_count += 1

    if not mapped_value_count:
        raise ValueError("Không tìm thấy cột tương thích giữa template nguồn và blank template shop đích.")

    if target_brand:
        for row_offset in range(len(source_rows)):
            destination_row_index = target_start_row + row_offset
            set_optional_value(
                destination_sheet,
                destination_row_index,
                destination_columns,
                target_brand,
                prefix="brand[",
            )
            set_optional_value(
                destination_sheet,
                destination_row_index,
                destination_columns,
                target_brand,
                prefix="manufacturer[",
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    destination_workbook.save(output_path)
    return {
        "source_columns": len(source_columns),
        "destination_columns": len(destination_columns),
        "mapped_columns": len(mapped_headers),
        "source_values": source_value_count,
        "mapped_values": mapped_value_count,
        "unmapped_columns": sorted(unmapped_headers)[:50],
    }


def find_column(columns: dict[str, int], exact: str | None = None, prefix: str | None = None) -> int:
    if exact and exact in columns:
        return columns[exact]
    if prefix:
        for name, index in columns.items():
            if name.startswith(prefix):
                return index
    raise ValueError(f"Template thiếu cột bắt buộc: {exact or prefix}")


def optional_column(columns: dict[str, int], prefix: str) -> int | None:
    for name, index in columns.items():
        if name.startswith(prefix):
            return index
    return None


def set_optional_value(
    worksheet: Any,
    row_index: int,
    columns: dict[str, int],
    value: Any,
    prefix: str,
) -> None:
    column_index = optional_column(columns, prefix)
    if column_index is not None:
        worksheet.cell(row_index, column_index).value = value


def set_image_locator_value(
    worksheet: Any,
    row_index: int,
    columns: dict[str, int],
    value: str,
    prefix: str,
    required: bool = False,
) -> None:
    column_index = (
        find_column(columns, prefix=prefix)
        if required
        else optional_column(columns, prefix)
    )
    if column_index is None:
        return
    cell = worksheet.cell(row_index, column_index)
    cell.value = value
    cell.hyperlink = value or None


def set_value(
    worksheet: Any,
    row_index: int,
    columns: dict[str, int],
    value: Any,
    exact: str | None = None,
    prefix: str | None = None,
) -> None:
    worksheet.cell(row_index, find_column(columns, exact=exact, prefix=prefix)).value = value


def set_repeated_value(
    worksheet: Any,
    row_index: int,
    columns: dict[str, int],
    prefix: str,
    occurrence: int,
    value: Any,
) -> None:
    marker = f"#{occurrence}."
    for name, column_index in columns.items():
        if name.startswith(prefix) and marker in name:
            worksheet.cell(row_index, column_index).value = value
            return
    raise ValueError(f"Template thiếu cột bắt buộc: {prefix} #{occurrence}")


def listing_text(item: dict[str, Any], key: str, fallback: Any = "") -> Any:
    return item.get("listing", {}).get(key, fallback)


def find_image_column(columns: dict[str, int], image_type: str, index: int = 1) -> int | None:
    if image_type == "main":
        patterns = ["main_product_image_locator", "main_image_url", "main_image"]
        for p in patterns:
            for name, col_idx in columns.items():
                if name.startswith(p):
                    return col_idx
        return None
    elif image_type == "swatch":
        patterns = ["swatch_product_image_locator", "swatch_image_url"]
        for p in patterns:
            for name, col_idx in columns.items():
                if name.startswith(p):
                    return col_idx
        return None
    elif image_type == "other":
        patterns = [
            f"other_product_image_locator_{index}",
            f"other_product_image_locator[#{index}.",
            f"other_product_image_locator{index}",
            f"other_image_url{index}",
            f"other_image_url_{index}",
            f"other_image_locator_{index}",
        ]
        for p in patterns:
            for name, col_idx in columns.items():
                if name.startswith(p):
                    return col_idx
        return None
    return None


def fill_listing_fields(
    worksheet: Any,
    row_index: int,
    columns: dict[str, int],
    item: dict[str, Any],
) -> None:
    set_value(worksheet, row_index, columns, listing_text(item, "title"), prefix="item_name")
    set_value(
        worksheet,
        row_index,
        columns,
        listing_text(item, "description"),
        prefix="product_description",
    )
    bullets = list(listing_text(item, "bullet_points", []))[:5]
    bullets.extend([""] * (5 - len(bullets)))
    for bullet_index, bullet in enumerate(bullets, start=1):
        set_repeated_value(
            worksheet,
            row_index,
            columns,
            "bullet_point[",
            bullet_index,
            bullet,
        )
    set_value(
        worksheet,
        row_index,
        columns,
        listing_text(item, "backend_search_terms"),
        prefix="generic_keyword",
    )

    main_col = find_image_column(columns, "main")
    if main_col:
        worksheet.cell(row_index, main_col).value = ""
        worksheet.cell(row_index, main_col).hyperlink = None

    swatch_col = find_image_column(columns, "swatch")
    if swatch_col:
        worksheet.cell(row_index, swatch_col).value = ""
        worksheet.cell(row_index, swatch_col).hyperlink = None

    for i in range(1, 9):
        other_col = find_image_column(columns, "other", i)
        if other_col:
            worksheet.cell(row_index, other_col).value = ""
            worksheet.cell(row_index, other_col).hyperlink = None

    image_urls = list(item.get("image_urls", []))[:7]
    if image_urls and main_col:
        worksheet.cell(row_index, main_col).value = image_urls[0]
        worksheet.cell(row_index, main_col).hyperlink = image_urls[0] or None

    for image_index, image_url in enumerate(image_urls[1:7], start=1):
        other_col = find_image_column(columns, "other", image_index)
        if other_col and image_url:
            worksheet.cell(row_index, other_col).value = image_url
            worksheet.cell(row_index, other_col).hyperlink = image_url or None


def replace_source_identifier(value: Any, source_sku: str, target_sku: str) -> Any:
    text = cell_text(value)
    if not text:
        return value
    if source_sku and source_sku in text:
        return text.replace(source_sku, target_sku)
    source_suffix = source_sku[-4:] if len(source_sku) >= 4 else source_sku
    target_suffix = target_sku[-4:] if len(target_sku) >= 4 else target_sku
    if source_suffix and text.endswith(source_suffix):
        return f"{text[:-len(source_suffix)]}{target_suffix}"
    return value


def sanitize_restored_row(
    worksheet: Any,
    row_index: int,
    source_skus: list[str],
    target_sku: str,
) -> None:
    for col_index in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row_index, col_index)
        text = cell_text(cell.value)
        if text:
            for s_sku in source_skus:
                if s_sku and s_sku in text:
                    text = text.replace(s_sku, target_sku)
            if "ABC123" in text:
                text = text.replace("ABC123", target_sku)
            cell.value = text


def clean_sku(raw: Any) -> str:
    s = cell_text(raw).strip()
    if not s:
        return "SKU"
    if "_" in s:
        parts = s.split("_")
        if len(parts) > 1 and " " in parts[1]:
            return parts[0].strip()
    elif " - " in s:
        return s.split(" - ")[0].strip()
    return s


def fill_template(template_path: Path, payload_path: Path, output_path: Path) -> None:
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    items = payload.get("items", [])
    if not items:
        raise ValueError("Không có listing thành công để xuất template.")
    keep_vba = template_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(template_path, keep_vba=keep_vba, data_only=False)
    worksheet = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
    attribute_row, _, data_row = configured_template_rows(worksheet)
    columns = technical_columns(worksheet, attribute_row)
    source_parent_row, source_child_row = source_template_rows(worksheet, columns, data_row)
    child_snapshot = snapshot_row(worksheet, source_child_row)
    child_height = worksheet.row_dimensions[source_child_row].height
    sku_column = find_column(columns, prefix="contribution_sku")
    source_parent_sku = cell_text(worksheet.cell(source_parent_row, sku_column).value)
    source_child_sku = cell_text(worksheet.cell(source_child_row, sku_column).value)
    model_column = optional_column(columns, "model_number")
    part_column = optional_column(columns, "part_number")
    color_column = optional_column(columns, "color[")
    source_child_model = worksheet.cell(source_child_row, model_column).value if model_column else None
    source_child_part = worksheet.cell(source_child_row, part_column).value if part_column else None
    source_child_color = worksheet.cell(source_child_row, color_column).value if color_column else None

    target_data_row = data_row or min(source_parent_row, source_child_row)
    if target_data_row <= attribute_row:
        raise ValueError("Dòng bắt đầu dữ liệu của template Amazon không hợp lệ.")
    for row_index in range(attribute_row + 1, worksheet.max_row + 1):
        for column_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row_index, column_index)
            if not isinstance(cell, MergedCell):
                cell.value = None
                cell.hyperlink = None
                cell.comment = None

    # Parent/child output is temporarily disabled. Every generated row is an
    # independent listing, even when the saved legacy template contains a pair.
    for offset, item in enumerate(items):
        row_index = target_data_row + offset
        sku = clean_sku(item.get("sku"))
        brand = cell_text(item.get("brand")) or "Generic"
        restore_row(worksheet, row_index, child_snapshot)
        sanitize_restored_row(worksheet, row_index, [source_parent_sku, source_child_sku], sku)
        worksheet.row_dimensions[row_index].height = child_height
        set_value(worksheet, row_index, columns, sku, prefix="contribution_sku")
        set_optional_value(worksheet, row_index, columns, "", prefix="parentage_level[")
        set_optional_value(worksheet, row_index, columns, "", prefix="child_parent_sku_relationship[")
        set_optional_value(worksheet, row_index, columns, "", prefix="variation_theme")
        set_optional_value(worksheet, row_index, columns, brand, prefix="brand[")
        set_optional_value(worksheet, row_index, columns, brand, prefix="manufacturer[")
        set_optional_value(worksheet, row_index, columns, replace_source_identifier(source_child_model, source_child_sku, sku), prefix="model_number")
        set_optional_value(worksheet, row_index, columns, replace_source_identifier(source_child_part, source_child_sku, sku), prefix="part_number")
        child_color = sku if cell_text(source_child_color).casefold() == source_child_sku.casefold() else source_child_color
        set_optional_value(worksheet, row_index, columns, child_color, prefix="color[")
        validate_shipping_for_row(workbook, worksheet, row_index, columns)
        fill_listing_fields(worksheet, row_index, columns, item)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def create_standard_listing_excel(payload_path: Path, output_path: Path) -> None:
    data = json.loads(payload_path.read_text(encoding="utf-8"))
    items = data.get("items", [])

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Amazon Listings"

    headers = [
        "seller-sku",
        "marketplace",
        "product-type",
        "item-name",
        "brand-name",
        "bullet-point1",
        "bullet-point2",
        "bullet-point3",
        "bullet-point4",
        "bullet-point5",
        "product-description",
        "generic-keywords",
        "main-image-url",
        "other-image-url1",
        "other-image-url2",
        "other-image-url3",
        "other-image-url4",
        "other-image-url5",
        "other-image-url6",
    ]
    worksheet.append(headers)

    for item in items:
        sku = item.get("sku", "")
        listing = item.get("listing", {})
        brand = item.get("brand", "")
        image_urls = list(item.get("image_urls", []))[:7]

        bullets = listing.get("bullet_points", [])
        while len(bullets) < 5:
            bullets.append("")

        row = [
            sku,
            "US",
            "Listing Desk",
            listing.get("title", ""),
            brand,
            bullets[0],
            bullets[1],
            bullets[2],
            bullets[3],
            bullets[4],
            listing.get("description", ""),
            listing.get("backend_search_terms", ""),
            image_urls[0] if len(image_urls) > 0 else "",
            image_urls[1] if len(image_urls) > 1 else "",
            image_urls[2] if len(image_urls) > 2 else "",
            image_urls[3] if len(image_urls) > 3 else "",
            image_urls[4] if len(image_urls) > 4 else "",
            image_urls[5] if len(image_urls) > 5 else "",
            image_urls[6] if len(image_urls) > 6 else "",
        ]
        worksheet.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

def resize_image_for_ai(input_path: Path, output_path: Path, max_dim: int = 1600, quality: int = 82) -> dict[str, Any]:
    from PIL import Image, ImageOps
    with Image.open(input_path) as img:
        img = ImageOps.exif_transpose(img)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        width, height = img.size
        if max(width, height) > max_dim:
            if width >= height:
                new_w = max_dim
                new_h = int(height * (max_dim / width))
            else:
                new_h = max_dim
                new_w = int(width * (max_dim / height))
            img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        img.save(output_path, format="JPEG", quality=quality, optimize=True)
    return {"status": "ok", "bytes": output_path.stat().st_size}


def main() -> None:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    parse_parser = subparsers.add_parser("parse")
    parse_parser.add_argument("source")

    inspect_parser = subparsers.add_parser("inspect")
    inspect_parser.add_argument("source")

    scan_parser = subparsers.add_parser("scan_template")
    scan_parser.add_argument("source")

    scan_fields_parser = subparsers.add_parser("scan_required_fields")
    scan_fields_parser.add_argument("source")
    scan_fields_parser.add_argument("payload", nargs="?", default=None)

    prepare_standalone_parser = subparsers.add_parser("prepare_standalone")
    prepare_standalone_parser.add_argument("source")
    prepare_standalone_parser.add_argument("payload")
    prepare_standalone_parser.add_argument("output")

    sample_parser = subparsers.add_parser("sample")
    sample_parser.add_argument("output")

    fill_parser = subparsers.add_parser("fill")
    fill_parser.add_argument("template")
    fill_parser.add_argument("payload")
    fill_parser.add_argument("output")

    map_template_parser = subparsers.add_parser("map_template")
    map_template_parser.add_argument("source")
    map_template_parser.add_argument("destination")
    map_template_parser.add_argument("output")
    map_template_parser.add_argument("--brand", default="")

    build_excel_parser = subparsers.add_parser("build_excel")
    build_excel_parser.add_argument("payload")
    build_excel_parser.add_argument("output")

    resize_parser = subparsers.add_parser("resize_image")
    resize_parser.add_argument("source")
    resize_parser.add_argument("output")
    resize_parser.add_argument("--max-dim", type=int, default=1600)
    resize_parser.add_argument("--quality", type=int, default=82)

    args = parser.parse_args()
    if args.command == "parse":
        print(json.dumps(parse_workbook(Path(args.source)), ensure_ascii=False))
    elif args.command == "inspect":
        print(json.dumps(inspect_template(Path(args.source)), ensure_ascii=False))
    elif args.command == "scan_template":
        print(json.dumps(scan_listing_template(Path(args.source)), ensure_ascii=False))
    elif args.command == "scan_required_fields":
        payload_file = Path(args.payload) if args.payload else None
        print(json.dumps(required_template_fields(Path(args.source), payload_file), ensure_ascii=False))
    elif args.command == "prepare_standalone":
        print(json.dumps(prepare_standalone_template(
            Path(args.source),
            Path(args.payload),
            Path(args.output),
        ), ensure_ascii=False))
    elif args.command == "sample":
        create_sample(Path(args.output))
    elif args.command == "fill":
        fill_template(Path(args.template), Path(args.payload), Path(args.output))
    elif args.command == "map_template":
        print(json.dumps(map_template_to_blank(
            Path(args.source),
            Path(args.destination),
            Path(args.output),
            args.brand,
        ), ensure_ascii=False))
    elif args.command == "build_excel":
        create_standard_listing_excel(Path(args.payload), Path(args.output))
    elif args.command == "resize_image":
        print(json.dumps(resize_image_for_ai(Path(args.source), Path(args.output), args.max_dim, args.quality), ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as error:  # noqa: BLE001
        print(str(error), file=sys.stderr)
        raise
