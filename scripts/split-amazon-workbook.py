#!/usr/bin/env python3
"""Create cumulative SKU test workbooks without rewriting an Amazon template."""

from __future__ import annotations

import argparse
import re
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", MAIN_NS)


def workbook_bytes(source: Path) -> tuple[str, bytes]:
    if source.suffix.lower() == ".zip":
        with ZipFile(source) as archive:
            workbooks = [
                name
                for name in archive.namelist()
                if Path(name).suffix.lower() in {".xlsx", ".xlsm"}
            ]
            if len(workbooks) != 1:
                raise ValueError("ZIP phải chứa đúng một workbook .xlsx hoặc .xlsm.")
            return Path(workbooks[0]).name, archive.read(workbooks[0])
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Nguồn phải là .zip, .xlsx hoặc .xlsm.")
    return source.name, source.read_bytes()


def template_sheet_path(data: bytes) -> str:
    with ZipFile(BytesIO(data)) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationship_targets = {
            relationship.attrib["Id"]: relationship.attrib["Target"]
            for relationship in relationships_root.findall(f"{{{PACKAGE_REL_NS}}}Relationship")
        }
        for sheet in workbook_root.findall(f"{{{MAIN_NS}}}sheets/{{{MAIN_NS}}}sheet"):
            if sheet.attrib.get("name") == "Template":
                relationship_id = sheet.attrib[f"{{{REL_NS}}}id"]
                target = relationship_targets[relationship_id].lstrip("/")
                return target if target.startswith("xl/") else f"xl/{target}"
    raise ValueError("Workbook không có sheet Template.")


def sku_rows(data: bytes) -> tuple[int, list[int]]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=False, keep_vba=True)
    worksheet = workbook["Template"]
    settings = str(worksheet.cell(1, 1).value or "")
    configured_attribute_row = re.search(r"(?:^|&)attributeRow=(\d+)(?:&|$)", settings)
    attribute_row = int(configured_attribute_row.group(1)) if configured_attribute_row else None
    if attribute_row is None:
        best = (0, 0)
        for row in range(1, min(worksheet.max_row, 100) + 1):
            values = [str(worksheet.cell(row, column).value or "") for column in range(1, worksheet.max_column + 1)]
            score = sum(any(value.startswith(marker) for value in values) for marker in ("contribution_sku", "product_type", "item_name["))
            if score > best[0]:
                best = (score, row)
        if best[0] < 2:
            raise ValueError("Không tìm thấy dòng technical headers trong template Amazon.")
        attribute_row = best[1]

    sku_column = None
    parentage_column = None
    for column in range(1, worksheet.max_column + 1):
        technical_name = str(worksheet.cell(attribute_row, column).value or "")
        if technical_name.startswith("contribution_sku"):
            sku_column = column
        elif technical_name.startswith("parentage_level["):
            parentage_column = column
    if sku_column is None:
        raise ValueError("Template thiếu cột contribution_sku.")

    parent_row = None
    children = []
    for row in range(attribute_row + 1, worksheet.max_row + 1):
        sku = str(worksheet.cell(row, sku_column).value or "").strip()
        if not sku:
            continue
        parentage = str(worksheet.cell(row, parentage_column).value or "") if parentage_column else ""
        if sku.upper().endswith("_MAIN") or parentage.casefold() == "parent":
            parent_row = row
        else:
            children.append(row)
    if parent_row is None or not children:
        raise ValueError("Không tìm thấy parent và child SKU trong template.")
    return parent_row, children


def clear_excluded_rows(data: bytes, excluded_rows: set[int]) -> bytes:
    sheet_path = template_sheet_path(data)
    source_buffer = BytesIO(data)
    output_buffer = BytesIO()
    with ZipFile(source_buffer) as source_archive, ZipFile(output_buffer, "w") as output_archive:
        for info in source_archive.infolist():
            content = source_archive.read(info.filename)
            if info.filename == sheet_path:
                root = ET.fromstring(content)
                sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
                if sheet_data is None:
                    raise ValueError("Sheet Template không có sheetData.")
                for row in sheet_data.findall(f"{{{MAIN_NS}}}row"):
                    if int(row.attrib.get("r", "0")) not in excluded_rows:
                        continue
                    for cell in row.findall(f"{{{MAIN_NS}}}c"):
                        for child in list(cell):
                            if child.tag in {
                                f"{{{MAIN_NS}}}v",
                                f"{{{MAIN_NS}}}is",
                                f"{{{MAIN_NS}}}f",
                            }:
                                cell.remove(child)
                        cell.attrib.pop("t", None)
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output_archive.writestr(info, content, compress_type=info.compress_type or ZIP_DEFLATED)
    return output_buffer.getvalue()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source")
    parser.add_argument("output_directory")
    parser.add_argument("--sizes", default="1,2,3")
    args = parser.parse_args()

    source = Path(args.source)
    output_directory = Path(args.output_directory)
    output_directory.mkdir(parents=True, exist_ok=True)
    original_name, data = workbook_bytes(source)
    parent_row, children = sku_rows(data)
    sizes = sorted({int(value) for value in args.sizes.split(",") if value.strip()})
    if not sizes or sizes[0] < 1 or sizes[-1] > len(children):
        raise ValueError(f"Số SKU test phải nằm trong khoảng 1-{len(children)}.")

    stem = Path(original_name).stem
    suffix = Path(original_name).suffix
    for size in sizes:
        keep_rows = {parent_row, *children[:size]}
        excluded_rows = set(children) - keep_rows
        output = clear_excluded_rows(data, excluded_rows)
        output_path = output_directory / f"{stem}_TEST_{size:02d}SKU{suffix}"
        output_path.write_bytes(output)
        print(output_path)


if __name__ == "__main__":
    main()
